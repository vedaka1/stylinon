from openpyxl import load_workbook
from src.domain.products.entities import Category, Product

files = {
    "cemtorg.xlsx": {
        "ПРАЙС_ЛИСТ_МОДИФИКАТ": {
            "total_columns": 10,
            "start_row": 5,
            "end_row": 53,
            "categories_row": (5, 17, 29, 39, 44, 51),
            "product_fields": {
                "weight": 2,
                "retail_price": 5,
                "wholesale_price": 6,
                "d1_delivery_price": 9,
                "d1_self_pickup_price": 10,
            },
        },
        "ПРАЙС ЛИСТ ЦПС Гарц": {
            "total_columns": 7,
            "start_row": 5,
            "end_row": 36,
            "categories_row": (5, 23, 27, 32),
            "product_fields": {
                "weight": 2,
                "retail_price": 5,
                "d1_delivery_price": 6,
                "d1_self_pickup_price": 7,
            },
        },
    },
}
result = {}
categories = []
for filename, sheets in files.items():
    wb = load_workbook(f"./products_data/{filename}")
    for key, value in sheets.items():
        sheet = wb[key]
        start_row = value["start_row"]
        end_row = value["end_row"]
        total_columns = value["total_columns"]
        categories_row = value["categories_row"]
        product_fields = value["product_fields"]
        category = None
        name_field = None

        for row in range(start_row, end_row):
            if row in categories_row:
                category = sheet.cell(row, 1).value
                result[category] = []
                categories.append(Category.create(name=category))
            else:
                product_data = {}

                if sheet.cell(row, 1).value is not None:
                    name_field = sheet.cell(row, 1).value

                product_data["name"] = name_field
                product_data["category"] = category
                product_data["sku"] = 123
                product_data["description"] = name_field

                for key, value in product_fields.items():
                    product_data[key] = sheet.cell(row, value).value

                product = Product.create(**product_data)
                result[category].append(product)
print(categories)
for data in result["ПЕСКОБЕТОН/СУХАЯ СМЕСЬ"]:
    print(data)
